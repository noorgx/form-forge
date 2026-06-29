"""
Batch generation: fill one template with many rows from an XLSX and render an
image per row.

Column convention (one row = one filled form):
  - text / digits field "name"        -> column "name"  (the value)
  - choice field "name"               -> column "name"  (the selected option key)
  - date field "name"                 -> columns "name.day", "name.month", "name.year"

Output filename: the value of a field whose name contains "serial" (else "row_N").
"""
import copy
import os
import re
import openpyxl

from . import renderer as form_filler

DATE_PARTS = ("day", "month", "year")


def fillable_columns(fields):
    """Ordered column headers for the fill spreadsheet."""
    cols = []
    for f in fields:
        name = f["name"]
        if f.get("type") == "date":
            cols += [f"{name}.{p}" for p in DATE_PARTS]
        else:  # text, digits, choice
            cols.append(name)
    return cols


def make_xlsx_template(fields, path):
    """Write a blank .xlsx with one header row of column names."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "fill"
    ws.append(fillable_columns(fields))
    wb.save(path)
    return path


def apply_row(fields, row):
    """Return a deep copy of `fields` with values filled from `row` (a dict of
    column -> cell value)."""
    out = copy.deepcopy(fields)
    for f in out:
        name = f["name"]
        t = f.get("type", "text")
        if t == "date":
            parts = f.setdefault("parts", {})
            for p in DATE_PARTS:
                v = row.get(f"{name}.{p}")
                if v not in (None, ""):
                    parts.setdefault(p, {})["value"] = str(v)
        elif t == "choice":
            v = row.get(name)
            if v not in (None, ""):
                f["selected"] = str(v)
        else:
            v = row.get(name)
            if v not in (None, ""):
                f["value"] = str(v)
    return out


def _safe(s):
    return re.sub(r"[^0-9A-Za-z._-]+", "_", s).strip("_") or "x"


def _filename_for(fields, row, idx):
    for f in fields:
        if "serial" in f["name"].lower():
            v = row.get(f["name"])
            if v not in (None, ""):
                return _safe(str(v))
    if row.get("serial") not in (None, ""):
        return _safe(str(row["serial"]))
    return f"row_{idx}"


def generate(meta, fields, xlsx_path, out_dir):
    """Render one image per data row. Returns list of output paths."""
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    made = []
    for idx, raw in enumerate(rows[1:], start=1):
        row = {headers[j]: raw[j] for j in range(min(len(headers), len(raw)))}
        if all(v in (None, "") for v in row.values()):
            continue  # skip blank rows
        flds = apply_row(fields, row)
        cfg = dict(meta)
        cfg["fields"] = flds
        cfg["out_image"] = os.path.join(out_dir, _filename_for(fields, row, idx) + ".png")
        form_filler.render_dict(cfg)
        made.append(cfg["out_image"])
    return made
